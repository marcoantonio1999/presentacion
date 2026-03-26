from __future__ import annotations

import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
BACKLOG_XLSX = Path(r"C:\Users\ordunama\Downloads\Tracking-Backlog-Claro-video.xlsx")
BACKLOG_SHEET_NAME = "backlog"
PHASE2_ZIP = Path(r"C:\Users\ordunama\Downloads\fase 2.zip")
VSS_ROOT = Path(r"C:\Users\ordunama\Downloads\vss consolidado")

OUTPUT_JSON = BASE_DIR / "ROADMAP_EJECUTIVO_PORTAFOLIO_20_REPOSITORIOS.json"
OUTPUT_JS = BASE_DIR / "ROADMAP_EJECUTIVO_PORTAFOLIO_20_REPOSITORIOS.data.js"

STACK_ORDER = ["Web", "Android", "Android TV", "iOS", "Roku", "Windows"]
THEME_ORDER = [
    "Pruebas unitarias",
    "Seguridad",
    "Rendimiento",
    "Observabilidad",
    "Calidad/Arquitectura",
    "Operación/Otras",
]
THEME_RANK = {theme: index + 1 for index, theme in enumerate(THEME_ORDER)}
DIMENSION_ORDER = ["Arquitectura", "Testing", "Seguridad", "DevOps", "Dependencias", "Trazabilidad"]
DIMENSION_THEME_MAP = {
    "Arquitectura": "Calidad/Arquitectura",
    "Testing": "Pruebas unitarias",
    "Seguridad": "Seguridad",
    "DevOps": "Operación/Otras",
    "Dependencias": "Calidad/Arquitectura",
    "Trazabilidad": "Observabilidad",
}

REPO_STACK = {
    "claro-video-web": "Web",
    "claro-video-web-chromecast": "Web",
    "claro-video-middleware-claro-musica": "Web",
    "claro-video-aaf-iptv": "Web",
    "claro-video-aaf-ott": "Web",
    "claro-video-aaf-lg-nativa": "Web",
    "claro-video-aaf-samsung-nativa": "Web",
    "claro-video-android": "Android",
    "claro-video-android-request-manager": "Android",
    "claro-video-android-coship": "Android",
    "claro-video-android-tv": "Android TV",
    "claro-video-android-tv-stb": "Android TV",
    "claro-video-android-smart-tv": "Android TV",
    "claro-video-ios": "iOS",
    "claro-video-ios-player": "iOS",
    "claro-video-ios-services": "iOS",
    "claro-video-ios-analytics": "iOS",
    "claro-video-ios-tvos": "iOS",
    "claro-video-roku-tv": "Roku",
    "claro-video-universal-windows-platform": "Windows",
}

REPO_ALIASES = {
    "claro-video-web-chromecast": ["claro-video-web-chromecast", "claro-video-chromecast"],
    "claro-video-aaf-lg-nativa": ["claro-video-aaf-lg-nativa", "claro-video-aff-lg-nativa", "claro-video-aaf-lg-nativa"],
    "claro-video-ios-services": ["claro-video-ios-services", "claro-video-services"],
}
TRACKING_SHEET_REPO = {
    "android-coship": "claro-video-android-coship",
    "android": "claro-video-android",
    "web-chromecast": "claro-video-web-chromecast",
    "web": "claro-video-web",
    "uwp": "claro-video-universal-windows-platform",
    "roku-tv": "claro-video-roku-tv",
    "ios-tvos": "claro-video-ios-tvos",
    "ios-player": "claro-video-ios-player",
    "ios-analytics": "claro-video-ios-analytics",
    "ios": "claro-video-ios",
    "android-tv-stb": "claro-video-android-tv-stb",
    "android-tv": "claro-video-android-tv",
    "android-smart-tv": "claro-video-android-smart-tv",
    "aaf-samsung-nativa": "claro-video-aaf-samsung-nativa",
    "aaf-ott": "claro-video-aaf-ott",
    "aaf-iptv": "claro-video-aaf-iptv",
    "aaf-lg-nativa": "claro-video-aaf-lg-nativa",
}

THEME_KEYWORDS = {
    "Pruebas unitarias": [
        "pruebas unitarias",
        "pruebas",
        "test",
        "tests",
        "testing",
        "unit test",
        "unitarias",
        "coverage",
        "cobertura",
        "mock",
        "mocks",
        "junit",
        "jest",
        "xctest",
        "assert",
        "pipeline de pruebas",
    ],
    "Seguridad": [
        "seguridad",
        "security",
        "credencial",
        "credenciales",
        "token",
        "jwt",
        "cve",
        "vulnerab",
        "https",
        "tls",
        "secreto",
        "secret",
        "password",
        "contrasena",
        "contraseña",
        "expuestas",
        "hardcode",
        "hardcoded",
        "auth",
        "autentic",
        "xss",
        "csrf",
        "inyeccion",
        "sql injection",
        "forjar",
        "drm",
        "sast",
    ],
    "Rendimiento": [
        "rendimiento",
        "performance",
        "latencia",
        "lento",
        "carga",
        "carga inicial",
        "tiempo de respuesta",
        "lazy loading",
        "bundle",
        "optimiz",
        "cache",
        "memoria",
        "cpu",
        "throughput",
        "escalabilidad",
        "bottleneck",
    ],
    "Observabilidad": [
        "observabilidad",
        "observability",
        "logging",
        "logs",
        "log",
        "monitor",
        "metrics",
        "métricas",
        "telemetría",
        "tracing",
        "alerta",
        "alerting",
        "visibilidad",
        "auditoria",
        "rastreo",
    ],
    "Calidad/Arquitectura": [
        "arquitect",
        "dependencia",
        "dependencias",
        "refactor",
        "refactoring",
        "manten",
        "mantenibilidad",
        "deuda técnica",
        "deuda tecnológica",
        "deuda tecnologica",
        "legacy",
        "obsolescencia",
        "god class",
        "duplic",
        "modular",
        "lint",
        "eslint",
        "tipado",
        "clean architecture",
        "documentación",
        "documentación",
        "portabilidad",
        "compatibilidad",
    ],
    "Operación/Otras": [
        "operacion",
        "operación",
        "ci/cd",
        "cicd",
        "pipeline",
        "release",
        "despliegue",
        "deploy",
        "soporte",
        "onboarding",
        "continuidad operativa",
        "runbook",
        "backup",
    ],
}

THEME_WINDOWS = {
    "Pruebas unitarias": ["S32", "S33"],
    "Seguridad": ["S34", "S35"],
    "Rendimiento": ["S36", "S37"],
    "Observabilidad": ["S38", "S39"],
    "Calidad/Arquitectura": ["S34", "S35", "S36", "S37", "S38", "S39", "S40"],
    "Operación/Otras": ["S41", "S42"],
}

PRIORITY_COLORS = {
    "Pruebas unitarias": "#14b8a6",
    "Seguridad": "#10b981",
    "Rendimiento": "#f59e0b",
    "Observabilidad": "#22d3ee",
    "Calidad/Arquitectura": "#84cc16",
    "Operación/Otras": "#94a3b8",
}
KICKOFF_DATE = date(2026, 3, 17)
SPRINT_CALENDAR = [
    ("S31", "Sprint 31", date(2026, 3, 17), date(2026, 3, 20)),
    ("S32", "Sprint 32", date(2026, 3, 23), date(2026, 4, 1)),
    ("S33", "Sprint 33", date(2026, 4, 6), date(2026, 4, 17)),
    ("S34", "Sprint 34", date(2026, 4, 20), date(2026, 4, 30)),
    ("S35", "Sprint 35", date(2026, 5, 4), date(2026, 5, 15)),
    ("S36", "Sprint 36", date(2026, 5, 18), date(2026, 5, 29)),
    ("S37", "Sprint 37", date(2026, 6, 1), date(2026, 6, 12)),
    ("S38", "Sprint 38", date(2026, 6, 15), date(2026, 6, 26)),
    ("S39", "Sprint 39", date(2026, 6, 29), date(2026, 7, 10)),
    ("S40", "Sprint 40", date(2026, 7, 13), date(2026, 7, 24)),
    ("S41", "Sprint 41", date(2026, 7, 27), date(2026, 8, 7)),
    ("S42", "Sprint 42", date(2026, 8, 10), date(2026, 8, 21)),
]
GENERAL_TRACK_COLORS = {
    "Quality Gate": "#38bdf8",
    "QA + IA": "#f97316",
    "Cierre formal": "#cbd5e1",
}
LEGACY_SPRINT_LABEL_RE = re.compile(r"\bS(1[01]|[1-9])\b")


@dataclass(frozen=True)
class SprintDefinition:
    sprint_id: str
    label: str
    start: date
    end: date


def build_sprints() -> list[SprintDefinition]:
    return [
        SprintDefinition(
            sprint_id=sprint_id,
            label=label,
            start=sprint_start,
            end=sprint_end,
        )
        for sprint_id, label, sprint_start, sprint_end in SPRINT_CALENDAR
    ]


SPRINTS = build_sprints()
SPRINT_BY_ID = {item.sprint_id: item for item in SPRINTS}


def shift_legacy_sprint_id(sprint_id: str) -> str:
    match = re.fullmatch(r"S(\d+)", sprint_id)
    if not match:
        return sprint_id
    value = int(match.group(1))
    if 1 <= value <= 11:
        return f"S{value + 31}"
    return sprint_id


def replace_legacy_sprint_tokens(text: str) -> str:
    return LEGACY_SPRINT_LABEL_RE.sub(lambda match: shift_legacy_sprint_id(match.group(0)), text)


def make_general_segment(
    *,
    title: str,
    theme: str,
    start_sprint: str,
    end_sprint: str,
    summary: str,
    hallazgos: str = "",
    activity: str = "",
    qa_ia: str = "",
    label: str | None = None,
    color: str | None = None,
) -> dict[str, object]:
    today = date.today()
    start = SPRINT_BY_ID[start_sprint]
    end = SPRINT_BY_ID[end_sprint]
    return {
        "title": title,
        "theme": theme,
        "label": label or theme,
        "color": color or PRIORITY_COLORS.get(theme, "#64748b"),
        "summary": summary,
        "hallazgos": hallazgos,
        "activity": activity,
        "qa_ia": qa_ia,
        "start_sprint": start_sprint,
        "end_sprint": end_sprint,
        "start": start.start.isoformat(),
        "end": end.end.isoformat(),
        "is_active": start.start <= today <= end.end,
        "duration_sprints": int(end_sprint[1:]) - int(start_sprint[1:]) + 1,
    }


def build_general_plan() -> dict[str, object]:
    legacy_qa_focus = {
        "S1": ("Flujos críticos base y criterios de aceptación.", "Diseño de smoke inicial y plan de pruebas.", "QA arma la suite base; IA acelera casos, datos y trazabilidad."),
        "S2": ("Defectos tempranos y estabilidad de regresión.", "Validación continua y afinación de cobertura.", "QA ejecuta regresión; IA apoya análisis de fallas y casos faltantes."),
        "S3": ("Hallazgos priorizados para remediación estructural.", "Validación funcional del cambio y de criterios del gate.", "QA corre pruebas de impacto; IA consolida evidencias y checklist."),
        "S4": ("Desviaciones sobre brechas críticas y deuda prioritaria.", "Pruebas de regresión y smoke sobre remediaciones.", "QA valida estabilidad; IA ayuda a comparar hallazgos abiertos vs cerrados."),
        "S5": ("Hallazgos de rendimiento y comportamiento degradado.", "Testeo focalizado de performance funcional y playback.", "QA mide regresión operativa; IA apoya lectura de patrones y reportes."),
        "S6": ("Pendientes de estabilidad y findings de segunda ola.", "Validación de cierre parcial y pruebas de no regresión.", "QA ejecuta suite extendida; IA propone casos derivados de defectos reales."),
        "S7": ("Brechas de monitoreo y observabilidad detectadas.", "Pruebas de visibilidad operativa y alertamiento.", "QA valida eventos y logs; IA ayuda a revisar cobertura de señales."),
        "S8": ("Hallazgos residuales de telemetría y soporte.", "Regresión completa de los cambios transversales.", "QA corre ronda ampliada; IA apoya triage y consolidado de evidencia."),
        "S9": ("Findings estructurales remanentes y deuda de cierre.", "Validación de arquitectura estabilizada y readiness.", "QA verifica regresión final de cambios grandes; IA acelera checklist final."),
        "S10": ("Incidencias previas a entrega y quality gate de salida.", "Pruebas integrales de pre-cierre y aprobación.", "QA ejecuta validación de salida; IA consolida evidencia y pendientes finales."),
        "S11": ("Pendientes mínimos y hallazgos de aceptación final.", "Validación integral, acta de cierre y handoff.", "QA realiza aceptación final; IA ayuda a empaquetar entregables y trazabilidad."),
    }
    qa_focus = {
        "S31": (
            "Arranque operativo, alcance inicial y criterios base de aceptación.",
            "Kickoff del programa, setup de validación y preparación del frente de pruebas.",
            "QA define cobertura inicial; IA apoya el armado de casos, datos y trazabilidad.",
        ),
        **{shift_legacy_sprint_id(sprint_id): values for sprint_id, values in legacy_qa_focus.items()},
    }

    legacy_quality_architecture = {
        "S3": ("S3 · Baseline de arquitectura", "El trabajo estructural deja de vivir solo en Sprint 9 y arranca de forma incremental.", "Hotspots, deuda técnica y findings base por repositorio.", "Priorización y primera ronda de remediación estructural.", "QA valida impacto inicial; IA acelera triage y documentación de cambios."),
        "S4": ("S4 · Remediación tranche 1", "Se ataca deuda prioritaria y acoplamientos críticos sobre los repositorios más sensibles.", "Duplicidad, dependencias rígidas y componentes de alto mantenimiento.", "Desacoplamiento inicial y limpieza de deuda prioritaria.", "QA corre regresión focalizada; IA compara hallazgos abiertos vs corregidos."),
        "S5": ("S5 · Remediación tranche 2", "La remediación estructural continúa junto con el frente de rendimiento.", "Hallazgos de modularidad y puntos de fallo recurrentes.", "Refactor incremental y estabilización de piezas compartidas.", "QA valida flujos recuperados; IA apoya análisis de hotspots residuales."),
        "S6": ("S6 · Consolidación técnica", "Se consolida la segunda ola de ajustes sobre componentes ya tocados.", "Hallazgos de mantenibilidad y deuda residual post-refactor.", "Normalización de patrones, limpieza y cierre parcial de deuda.", "QA ejecuta regresión extendida; IA ayuda a priorizar remanentes."),
        "S7": ("S7 · Continuidad de remediación", "El frente estructural se mantiene mientras entra observabilidad.", "Hallazgos que afectan soporte, trazabilidad y continuidad operativa.", "Ajustes estructurales ligados a logs, eventos y consistencia.", "QA valida escenarios de monitoreo; IA cruza hallazgos con evidencia operativa."),
        "S8": ("S8 · Cierre de hallazgos mayores", "Se enfoca el esfuerzo en los findings de mayor costo de mantenimiento.", "Hallazgos mayores aún abiertos y deuda que bloquea estabilización.", "Cierre de hallazgos grandes y estabilización final del frente.", "QA confirma no regresión; IA consolida cierre de hallazgos y criterios pendientes."),
        "S9": ("S9 · Ajuste estructural final", "Sprint de cierre técnico del frente distribuido desde S3.", "Hallazgos residuales y puntos finos previos al gate de salida.", "Normalización final, cleanup y preparación para pre-cierre.", "QA valida readiness estructural; IA empaqueta trazabilidad y evidencia."),
    }
    quality_architecture = {
        shift_legacy_sprint_id(sprint_id): tuple(replace_legacy_sprint_tokens(part) for part in values)
        for sprint_id, values in legacy_quality_architecture.items()
    }

    legacy_quality_gate = {
        "S3": ("S3 · Gate de entrada", "Arranca el nuevo trabajo de Quality Gate con criterios mínimos transversales.", "Hallazgos priorizados que abren el gate inicial.", "Definición del gate, checklist base y umbrales de aceptación.", "QA valida criterios; IA ayuda a consolidar evidencia y checklist."),
        "S4": ("S4 · Gate sobre cambios críticos", "El gate acompaña la primera ola de remediación.", "Desviaciones sobre cambios críticos y brechas repetitivas.", "Aplicación de criterios de aprobación sobre cambios sensibles.", "QA ejecuta el gate; IA resume resultados y riesgos residuales."),
        "S5": ("S5 · Gate de rendimiento", "Se incorporan criterios ligados a estabilidad y performance.", "Hallazgos de rendimiento con impacto en experiencia.", "Revisiones de salida sobre cambios de performance.", "QA corre validación funcional; IA ayuda a contrastar findings y evidencia."),
        "S6": ("S6 · Gate de continuidad", "Se consolida el gate como práctica recurrente del programa.", "Hallazgos reincidentes y deuda aún abierta.", "Seguimiento de cumplimiento y excepciones controladas.", "QA valida el cumplimiento; IA asiste con trazabilidad y resumen ejecutivo."),
        "S7": ("S7 · Gate de observabilidad", "El gate cubre ahora criterios de monitoreo y señales operativas.", "Hallazgos de eventos, logs y alertamiento incompleto.", "Verificación de criterios de visibilidad operativa.", "QA valida telemetría; IA ayuda a revisar consistencia de eventos y evidencia."),
        "S8": ("S8 · Gate ampliado", "Se amplian los criterios de aceptación previos al cierre.", "Hallazgos residuales transversales y excepciones abiertas.", "Control de calidad previo a estabilización final.", "QA ejecuta regression gate; IA consolida pendientes y decisiones."),
        "S9": ("S9 · Gate sobre cierre estructural", "El frente estructural ya se valida contra un gate formal.", "Hallazgos que aún condicionan readiness técnico.", "Aprobacion controlada de cierres y deuda remanente.", "QA valida readiness; IA apoya checklist y evidencia de aprobación."),
        "S10": ("S10 · Gate de salida", "Último quality gate antes del cierre formal del proyecto.", "Pendientes finales, excepciones y riesgos de entrega.", "Gate previo a entrega, congelamiento y handoff.", "QA ejecuta validación de salida; IA consolida paquete final de evidencia."),
    }
    quality_gate = {
        shift_legacy_sprint_id(sprint_id): tuple(replace_legacy_sprint_tokens(part) for part in values)
        for sprint_id, values in legacy_quality_gate.items()
    }

    rows = [
        {
            "lane": "Pruebas unitarias",
            "status": "Arranque S31 y trabajo inicial S32-S33",
            "segments": [
                make_general_segment(
                    title="S31 · Arranque operativo",
                    theme="Pruebas unitarias",
                    start_sprint="S31",
                    end_sprint="S31",
                    summary="Semana corta para arrancar el programa, fijar criterios y preparar el frente inicial.",
                    hallazgos="Cobertura base por confirmar en flujos críticos y huecos iniciales por validar.",
                    activity="Kickoff del roadmap, setup base, alineación técnica y definición de primeros casos.",
                    qa_ia="QA define cobertura inicial; IA acelera preparación de casos, datos y trazabilidad.",
                ),
                make_general_segment(
                    title="S32 · Baseline y primeros casos",
                    theme="Pruebas unitarias",
                    start_sprint="S32",
                    end_sprint="S32",
                    summary="Se formaliza y amplía el arranque iniciado en S31 sobre los primeros repositorios objetivo.",
                    hallazgos="Cobertura base ausente o parcial en flujos críticos.",
                    activity="Setup de harnesses, mocks, smoke inicial y primeros casos unitarios.",
                    qa_ia="QA arma la suite base; IA acelera casos, datos y trazabilidad.",
                ),
                make_general_segment(
                    title="S33 · Estabilización y expansión",
                    theme="Pruebas unitarias",
                    start_sprint="S33",
                    end_sprint="S33",
                    summary="Segunda ola para estabilizar el frente inicial y ampliar cobertura.",
                    hallazgos="Defectos tempranos y huecos de regresión detectados en la primera ronda.",
                    activity="Ampliación de suites críticas, smoke reforzado y estabilización de regresión.",
                    qa_ia="QA ejecuta regresión continua; IA ayuda a priorizar casos faltantes y defectos.",
                ),
            ],
        },
        {
            "lane": "QA + IA transversal",
            "status": "Actividad adicional en todos los sprints S31-S42",
            "segments": [
                make_general_segment(
                    title=f"{sprint_id} · Validación y testeo",
                    theme="Observabilidad",
                    label="QA + IA",
                    color=GENERAL_TRACK_COLORS["QA + IA"],
                    start_sprint=sprint_id,
                    end_sprint=sprint_id,
                    summary="Frente transversal de acompañamiento, validación, evidencia y soporte de calidad.",
                    hallazgos=hallazgos,
                    activity=activity,
                    qa_ia=qa_ia,
                )
                for sprint_id, (hallazgos, activity, qa_ia) in qa_focus.items()
            ],
        },
        {
            "lane": "Seguridad",
            "status": "Brechas críticas S34-S35",
            "segments": [
                make_general_segment(
                    title="Seguridad · CVEs, secretos y autenticación",
                    theme="Seguridad",
                    start_sprint="S34",
                    end_sprint="S35",
                    summary="Se corrigen exposiciones de mayor impacto mientras avanza la remediación estructural.",
                    hallazgos="CVEs, secretos, hardcodes y debilidades de autenticación priorizadas.",
                    activity="Corrección de brechas críticas, blindaje de datos y endurecimiento de accesos.",
                    qa_ia="QA valida escenarios sensibles; IA ayuda a consolidar hallazgos y evidencia de cierre.",
                ),
            ],
        },
        {
            "lane": "Calidad / Arquitectura",
            "status": "Distribuido de S34 a S40",
            "segments": [
                make_general_segment(
                    title=title,
                    theme="Calidad/Arquitectura",
                    start_sprint=sprint_id,
                    end_sprint=sprint_id,
                    summary=summary,
                    hallazgos=hallazgos,
                    activity=activity,
                    qa_ia=qa_ia,
                )
                for sprint_id, (title, summary, hallazgos, activity, qa_ia) in quality_architecture.items()
            ],
        },
        {
            "lane": "Quality Gate",
            "status": "Nuevo trabajo de S34 a S41",
            "segments": [
                make_general_segment(
                    title=title,
                    theme="Seguridad",
                    label="Quality Gate",
                    color=GENERAL_TRACK_COLORS["Quality Gate"],
                    start_sprint=sprint_id,
                    end_sprint=sprint_id,
                    summary=summary,
                    hallazgos=hallazgos,
                    activity=activity,
                    qa_ia=qa_ia,
                )
                for sprint_id, (title, summary, hallazgos, activity, qa_ia) in quality_gate.items()
            ],
        },
        {
            "lane": "Rendimiento",
            "status": "Optimización S36-S37",
            "segments": [
                make_general_segment(
                    title="Rendimiento · Performance y estabilidad",
                    theme="Rendimiento",
                    start_sprint="S36",
                    end_sprint="S37",
                    summary="Bloque dedicado a performance, playback, cargas y respuesta de componentes sensibles.",
                    hallazgos="Cuellos de botella, latencias y degradaciones con impacto directo al usuario.",
                    activity="Optimización de cargas, playback, bundles y tiempos de respuesta.",
                    qa_ia="QA ejecuta pruebas focalizadas; IA apoya lectura de patrones y comparativos de hallazgos.",
                ),
            ],
        },
        {
            "lane": "Observabilidad",
            "status": "Visibilidad S38-S39",
            "segments": [
                make_general_segment(
                    title="Observabilidad · Logs, métricas y alertas",
                    theme="Observabilidad",
                    start_sprint="S38",
                    end_sprint="S39",
                    summary="Frente transversal para elevar la visibilidad operativa antes del pre-cierre.",
                    hallazgos="Brechas en logs, eventos, telemetría y alertamiento operacional.",
                    activity="Normalización de eventos, señalización y trazabilidad operativa.",
                    qa_ia="QA valida evidencia operativa; IA apoya el cruce de señales y checklist.",
                ),
            ],
        },
        {
            "lane": "Entrega y cierre",
            "status": "Pre-cierre S41 y cierre formal S42",
            "segments": [
                make_general_segment(
                    title="S41 · Pre-cierre y handoff",
                    theme="Operación/Otras",
                    start_sprint="S41",
                    end_sprint="S41",
                    summary="Sprint de preparación de entrega, estabilización final y handoff controlado.",
                    hallazgos="Pendientes finales, excepciones y observaciones previas al cierre formal.",
                    activity="Consolidación de backlog residual, paquete de entrega y readiness del proyecto.",
                    qa_ia="QA ejecuta validación pre-cierre; IA apoya paquete final de evidencia y seguimiento.",
                ),
                make_general_segment(
                    title="S42 · Cierre formal del proyecto",
                    theme="Operación/Otras",
                    label="Cierre formal",
                    color=GENERAL_TRACK_COLORS["Cierre formal"],
                    start_sprint="S42",
                    end_sprint="S42",
                    summary="Sprint dedicado al cierre formal, validación final, entrega integral y confirmación de que todo este OK.",
                    hallazgos="Últimos hallazgos de aceptación, pendientes mínimos y validaciones finales.",
                    activity="Cierre formal, entrega final, acta de aceptación y validación de entregables.",
                    qa_ia="QA realiza la aprobación final; IA ayuda a consolidar trazabilidad, checklists y paquete de cierre.",
                ),
            ],
        },
    ]

    return {
        "kickoff": KICKOFF_DATE.isoformat(),
        "official_start": SPRINT_BY_ID["S32"].start.isoformat(),
        "note": "El arranque operativo ocurre en S31 desde el 17-mar-2026 y los sprints formales van de S32 a S42 del 23-mar-2026 al 21-ago-2026, con cierres ajustados por Semana Santa y por el asueto del 01-may-2026.",
        "rows": rows,
    }


def strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    )


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = strip_accents(text.lower())
    text = re.sub(r"\s+", " ", text)
    return text


def slugify(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "sin-valor"


def normalize_repo(repo: object) -> str:
    text = str(repo or "").strip()
    if text == "claro-video-services":
        return "claro-video-ios-services"
    return text


def normalize_criticality(raw: object) -> tuple[str, int]:
    text = normalize_text(raw)
    if "critic" in text:
        return "CRITICA", 4
    if "alta" in text:
        return "ALTA", 3
    if "media" in text:
        return "MEDIA", 2
    if "baja" in text:
        return "BAJA", 1
    return "MEDIA", 2


def normalize_sprint_hint(raw_sprint: object, increment: object, period: object) -> str | None:
    text = " ".join([normalize_text(raw_sprint), normalize_text(increment), normalize_text(period)]).strip()
    if not text:
        return None
    if "semana 1" in text or "kickoff" in text:
        return "S31"
    if "incremento 0" in text or "sprint p0" in text or "sprint 0" in text:
        return "S32"
    if "incremento 1" in text or "sprint p1" in text or "semana 3-4" in text:
        return "S33"
    numbers = [int(number) for number in re.findall(r"\b(\d+)\b", text)]
    if not numbers:
        return None
    value = numbers[0]
    if 31 <= value <= 42:
        return f"S{value}"
    if value <= 0:
        return "S32"
    if value > 11:
        value = 11
    # Excel mixes zero-based increments and explicit sprint numbering.
    if "incremento" in text and value >= 0:
        value += 1
    if value > 11:
        value = 11
    return shift_legacy_sprint_id(f"S{value}")


def normalize_dimension(raw: object) -> str | None:
    text = normalize_text(raw)
    mapping = {
        "arquitectura": "Arquitectura",
        "testing": "Testing",
        "seguridad": "Seguridad",
        "devops": "DevOps",
        "dependencias": "Dependencias",
        "trazabilidad": "Trazabilidad",
    }
    return mapping.get(text)


def build_theme_scores(theme: str) -> dict[str, int]:
    return {candidate: 10 if candidate == theme else 0 for candidate in THEME_ORDER}


def row_value(raw_row: tuple[object, ...], header_map: dict[str, int], key: str) -> object:
    index = header_map.get(key)
    if index is None:
        return None
    return raw_row[index]


def classify_theme(record_text: str) -> tuple[str, str, bool, dict[str, int]]:
    scores: dict[str, int] = {theme: 0 for theme in THEME_ORDER}
    for theme, keywords in THEME_KEYWORDS.items():
        for keyword in keywords:
            if keyword in record_text:
                scores[theme] += 3 if " " in keyword else 2

    if any(token in record_text for token in ("riesgo de cumplimiento", "exposicion de datos", "exposicion de credenciales")):
        scores["Seguridad"] += 3
    if any(token in record_text for token in ("bajo nivel de deteccion", "visibilidad de calidad", "telemetría")):
        scores["Observabilidad"] += 3
    if any(token in record_text for token in ("obsolescencia", "bloqueo tecnologico", "mantenibilidad", "refactor")):
        scores["Calidad/Arquitectura"] += 3
    if any(token in record_text for token in ("escalabilidad", "tiempo de respuesta")):
        scores["Rendimiento"] += 3

    best_theme = max(THEME_ORDER, key=lambda item: (scores[item], -THEME_RANK[item]))
    best_score = scores[best_theme]
    if best_score >= 7:
        return best_theme, "alta", False, scores
    if best_score >= 4:
        return best_theme, "media", False, scores
    if best_score >= 2:
        return best_theme, "baja", best_theme == "Calidad/Arquitectura", scores
    return "Calidad/Arquitectura", "baja", True, scores


def infer_priority_theme(row: dict[str, object], record_text: str) -> tuple[str, str, bool, dict[str, int], str | None]:
    dimension = normalize_dimension(row.get("Dimensión"))
    heuristic_theme, heuristic_confidence, heuristic_review, heuristic_scores = classify_theme(record_text)
    if not dimension:
        return heuristic_theme, heuristic_confidence, heuristic_review, heuristic_scores, None

    mapped_theme = DIMENSION_THEME_MAP[dimension]
    if mapped_theme in {"Calidad/Arquitectura", "Operación/Otras"} and heuristic_theme in {"Rendimiento", "Observabilidad"}:
        return heuristic_theme, heuristic_confidence, False, heuristic_scores, dimension
    return mapped_theme, "alta", False, build_theme_scores(mapped_theme), dimension


def alias_candidates(repo: str) -> list[str]:
    candidates = [repo]
    candidates.extend(REPO_ALIASES.get(repo, []))
    # Avoid duplicates while preserving order.
    ordered: list[str] = []
    for candidate in candidates:
        if candidate not in ordered:
            ordered.append(candidate)
    return ordered


def find_vss_source(repo: str) -> str | None:
    for candidate in alias_candidates(repo):
        candidate_path = VSS_ROOT / candidate
        if candidate_path.exists():
            return str(candidate_path)
    return None


def build_phase2_index() -> dict[str, list[str]]:
    index: dict[str, list[str]] = defaultdict(list)
    if not PHASE2_ZIP.exists():
        return index
    with zipfile.ZipFile(PHASE2_ZIP) as archive:
        for entry in archive.namelist():
            if entry.endswith("/"):
                continue
            parts = entry.split("/")
            if len(parts) < 2:
                continue
            repo_folder = parts[0]
            for candidate in REPO_STACK:
                if repo_folder in alias_candidates(candidate):
                    index[candidate].append(entry)
    return index


PHASE2_INDEX = build_phase2_index()


def pick_phase2_refs(repo: str) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    annex = BASE_DIR / f"ANEXO_FASE2_{repo}.html"
    if annex.exists():
        refs.append({"label": "Anexo FASE 2", "href": annex.name})

    interesting = [
        "03-diagnostico/README.md",
        "03-diagnostico/TECHNICAL_DEBT.md",
        "03-diagnostico/SECURITY_REVIEW.md",
        "03-diagnostico/ISO_5055_QUALITY_REPORT.md",
        "diagrams/ARCHITECTURE_DIAGRAMS.md",
        "diagrams/ISO_5055_HEATMAP.md",
    ]
    candidates = PHASE2_INDEX.get(repo, [])
    for suffix in interesting:
        match = next((entry for entry in candidates if entry.endswith(suffix)), None)
        if match:
            refs.append({"label": suffix.split("/")[-1], "path": f"fase 2.zip::{match}"})
    return refs[:4]


def pick_vss_refs(repo: str) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    annex = BASE_DIR / f"ANEXO_{repo}.html"
    viewer = BASE_DIR / f"ANEXO_{repo}-diagramas-vss-viewer.html"
    if annex.exists():
        refs.append({"label": "Anexo VSS", "href": annex.name})
    if viewer.exists():
        refs.append({"label": "Diagramas VSS", "href": viewer.name})
    source = find_vss_source(repo)
    if source:
        refs.append({"label": "Fuente consolidada VSS", "path": source})
    return refs[:3]


def load_dimension_catalog(workbook) -> list[dict[str, str]]:
    if "Dimensiones" not in workbook.sheetnames:
        return []

    worksheet = workbook["Dimensiones"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []

    header_map = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    if "Dimensión" not in header_map:
        return []

    catalog: list[dict[str, str]] = []
    for raw_row in rows:
        if not any(cell not in (None, "") for cell in raw_row):
            continue
        dimension = normalize_dimension(row_value(raw_row, header_map, "Dimensión"))
        if not dimension:
            continue
        benefits = str(row_value(raw_row, header_map, "Beneficios Representativos") or "").strip()
        catalog.append({"dimension": dimension, "benefits": benefits})

    order_index = {dimension: index for index, dimension in enumerate(DIMENSION_ORDER)}
    catalog.sort(key=lambda item: order_index.get(item["dimension"], 99))
    return catalog


def infer_tracking_repo(sheet_name: str) -> str | None:
    normalized = normalize_text(sheet_name)
    prefix = re.sub(r"-?tracking.*$", "", normalized).strip("- ")
    return TRACKING_SHEET_REPO.get(prefix)


def normalize_tracking_status(raw: object) -> str:
    text = normalize_text(raw)
    if not text:
        return "SIN ESTADO"
    if "pend" in text:
        return "PENDIENTE"
    if "progreso" in text or "curso" in text:
        return "EN PROGRESO"
    if "bloq" in text:
        return "BLOQUEADO"
    if "complet" in text or "cerrad" in text or "done" in text:
        return "COMPLETADO"
    if "ninguno" in text:
        return "SIN ESTADO"
    return str(raw).strip().upper()


def normalize_tracking_sprint(raw: object) -> str | None:
    text = normalize_text(raw)
    if not text:
        return None

    explicit = re.search(r"\bs(3[1-9]|4[0-2])\b", text)
    if explicit:
        return explicit.group(0).upper()

    week_numbers = [int(number) for number in re.findall(r"\b(\d+)\b", text)]
    if "semana" in text and week_numbers:
        max_week = max(week_numbers)
        if max_week == 1:
            return "S31"
        if max_week == 2:
            return "S32"
        return f"S{min(42, 33 + ((max_week - 3) // 2))}"

    sprint_match = re.search(r"(?:sprint|incremento)\s*p?(\d+)", text)
    if sprint_match:
        value = int(sprint_match.group(1))
        return f"S{min(42, value + 32)}"

    return None


def load_tracking_actions(workbook) -> dict[str, object]:
    actions: list[dict[str, object]] = []
    used_sheets: list[str] = []

    for sheet_name in workbook.sheetnames:
        if "tracking" not in normalize_text(sheet_name):
            continue
        repo = infer_tracking_repo(sheet_name)
        if not repo or repo not in REPO_STACK:
            continue

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue
        header_map = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
        if "Estado" not in header_map or "Titulo" not in header_map:
            continue

        used_sheets.append(sheet_name)
        for raw_row in rows:
            if not any(cell not in (None, "") for cell in raw_row):
                continue

            title = str(row_value(raw_row, header_map, "Titulo") or "").strip()
            code = str(row_value(raw_row, header_map, "Codigo") or "").strip()
            if not title and not code:
                continue

            criticality_label, criticality_rank = normalize_criticality(row_value(raw_row, header_map, "Criticidad"))
            effort_min = coerce_int(row_value(raw_row, header_map, "Esfuerzo_Min_h"))
            effort_max = coerce_int(row_value(raw_row, header_map, "Esfuerzo_Max_h"))
            owner = str(row_value(raw_row, header_map, "Responsable") or "").strip() or "Sin responsable"
            sprint_raw = str(row_value(raw_row, header_map, "Sprint") or "").strip()
            actions.append(
                {
                    "sheet": sheet_name,
                    "repo": repo,
                    "stack": REPO_STACK[repo],
                    "code": code or str(row_value(raw_row, header_map, "ID") or "").strip(),
                    "block": str(row_value(raw_row, header_map, "Bloque") or "").strip(),
                    "area": str(row_value(raw_row, header_map, "Area") or "").strip(),
                    "title": title,
                    "description": str(row_value(raw_row, header_map, "Descripcion") or "").strip(),
                    "criticality": criticality_label,
                    "criticality_rank": criticality_rank,
                    "status": normalize_tracking_status(row_value(raw_row, header_map, "Estado")),
                    "sprint_raw": sprint_raw,
                    "sprint_assigned": normalize_tracking_sprint(row_value(raw_row, header_map, "Sprint")),
                    "effort_min_h": effort_min,
                    "effort_max_h": effort_max,
                    "effort_max_visible": max(effort_min, effort_max),
                    "remediation_type": str(row_value(raw_row, header_map, "Tipo_Remediacion") or "").strip(),
                    "dependencies": str(row_value(raw_row, header_map, "Dependencias") or "").strip(),
                    "implementation_risk": str(row_value(raw_row, header_map, "Riesgo_Implementacion") or "").strip(),
                    "owasp": str(row_value(raw_row, header_map, "OWASP") or "").strip(),
                    "owner": owner,
                    "notes": str(row_value(raw_row, header_map, "Notas") or "").strip(),
                    "source_path": f"{BACKLOG_XLSX.name}::{sheet_name}",
                }
            )

    actions.sort(
        key=lambda item: (
            STACK_ORDER.index(item["stack"]),
            sprintOrder_index(item["sprint_assigned"]),
            -item["criticality_rank"],
            item["repo"],
            item["code"],
        )
    )

    return {
        "source_path": str(BACKLOG_XLSX),
        "sheet_count": len(used_sheets),
        "sheet_names": sorted(used_sheets),
        "actions": actions,
    }


def value_stream_group(row: dict[str, object]) -> str:
    domain = str(row.get("Dominio_VSS") or "").strip()
    stream = str(row.get("Value_Stream") or "").strip()
    spec = str(row.get("Especificacion_VSS") or "").strip()
    if domain and domain.upper() not in {"N/A", "NA", "NONE"}:
        return domain
    if stream:
        return stream
    return spec or "Sin agrupar"


def coerce_int(value: object, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return default


def sprintOrder_index(sprint_id: str | None) -> int:
    if sprint_id and sprint_id in SPRINT_BY_ID:
        return sprintOrder_cache[sprint_id]
    return len(SPRINTS) + 1


sprintOrder_cache = {sprint.sprint_id: index for index, sprint in enumerate(SPRINTS)}


def sheet_records(workbook) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    worksheets = [workbook[BACKLOG_SHEET_NAME]] if BACKLOG_SHEET_NAME in workbook.sheetnames else workbook.worksheets
    for worksheet in worksheets:
        try:
            headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        header_map = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
        if "Repositorio" not in header_map:
            continue
        for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
            repo = normalize_repo(raw_row[header_map["Repositorio"]])
            if not repo or repo not in REPO_STACK:
                continue
            if not any(cell not in (None, "") for cell in raw_row):
                continue
            row = {header: raw_row[index] for header, index in header_map.items()}
            row["_sheet"] = worksheet.title
            rows.append(row)
    return rows


def assign_group_sprints(records: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    grouped: dict[tuple[str, str, str], list[dict[str, object]]] = defaultdict(list)
    for record in records:
        key = (record["stack"], record["priority_theme"], record["value_stream_group"])
        grouped[key].append(record)

    blocks: list[dict[str, object]] = []
    loads = {sprint.sprint_id: 0 for sprint in SPRINTS}

    for theme in THEME_ORDER:
        theme_blocks: list[dict[str, object]] = []
        for (stack, block_theme, group_name), bucket in grouped.items():
            if block_theme != theme:
                continue
            severity_score = sum(item["business_criticality_rank"] for item in bucket)
            critical_high = sum(1 for item in bucket if item["business_criticality_rank"] >= 3)
            repos = sorted({item["repo"] for item in bucket})
            theme_blocks.append(
                {
                    "stack": stack,
                    "theme": theme,
                    "value_stream_group": group_name,
                    "records": bucket,
                    "severity_score": severity_score,
                    "critical_high_count": critical_high,
                    "record_count": len(bucket),
                    "repos": repos,
                    "hint": Counter(item["sprint_hint"] for item in bucket if item["sprint_hint"]).most_common(1),
                }
            )

        theme_blocks.sort(
            key=lambda item: (
                -item["severity_score"],
                -item["critical_high_count"],
                -item["record_count"],
                STACK_ORDER.index(item["stack"]),
                item["value_stream_group"],
            )
        )

        allowed = THEME_WINDOWS[theme]
        theme_loads = {sprint_id: loads[sprint_id] for sprint_id in allowed}
        for block in theme_blocks:
            preferred = block["hint"][0][0] if block["hint"] else None
            if preferred not in allowed:
                preferred = None

            if preferred:
                chosen_sprint = min(
                    allowed,
                    key=lambda sprint_id: (
                        0 if sprint_id == preferred else 1,
                        theme_loads[sprint_id],
                        allowed.index(sprint_id),
                    ),
                )
            else:
                chosen_sprint = min(
                    allowed,
                    key=lambda sprint_id: (theme_loads[sprint_id], allowed.index(sprint_id)),
                )

            sprint = SPRINT_BY_ID[chosen_sprint]
            block["sprint_id"] = sprint.sprint_id
            block["sprint_label"] = sprint.label
            block["sprint_start"] = sprint.start.isoformat()
            block["sprint_end"] = sprint.end.isoformat()
            block["value_streams"] = sorted({record["value_stream"] for record in block["records"] if record["value_stream"]})
            block["sample_pain_points"] = [record["pain_point"] for record in block["records"][:3]]
            theme_loads[chosen_sprint] += block["record_count"]
            loads[chosen_sprint] += block["record_count"]
            blocks.append(block)

            for record in block["records"]:
                record["sprint_assigned"] = sprint.sprint_id
                record["sprint_start"] = sprint.start.isoformat()
                record["sprint_end"] = sprint.end.isoformat()

    blocks.sort(key=lambda item: (STACK_ORDER.index(item["stack"]), THEME_RANK[item["theme"]], item["sprint_id"], item["value_stream_group"]))
    return records, blocks


def build_dataset() -> dict[str, object]:
    workbook = load_workbook(BACKLOG_XLSX, data_only=True, read_only=True)
    try:
        rows = sheet_records(workbook)
        dimension_catalog = load_dimension_catalog(workbook)
        tracking = load_tracking_actions(workbook)
    finally:
        workbook.close()

    records: list[dict[str, object]] = []
    for index, row in enumerate(rows, start=1):
        repo = normalize_repo(row.get("Repositorio"))
        stack = REPO_STACK[repo]
        criticality_label, criticality_rank = normalize_criticality(row.get("Criticidad_Negocio"))
        text_blob = " ".join(
            normalize_text(row.get(key))
            for key in (
                "Pain_Point_Negocio",
                "Descripcion_Problema",
                "Estado_Actual",
                "Estado_Objetivo",
                "Beneficio",
                "Categoria_Riesgo",
                "Value_Stream",
                "Dominio_VSS",
                "Especificacion_VSS",
            )
        )
        theme, confidence, needs_review, theme_scores, dimension = infer_priority_theme(row, text_blob)
        source_sheet = row["_sheet"]
        raw_record = {
            "id": index,
            "repo": repo,
            "stack": stack,
            "channel": stack,
            "value_stream": str(row.get("Value_Stream") or "").strip() or "Sin value stream",
            "value_stream_group": value_stream_group(row),
            "domain_vss": str(row.get("Dominio_VSS") or "").strip(),
            "vss_spec": str(row.get("Especificacion_VSS") or "").strip(),
            "pain_point": str(row.get("Pain_Point_Negocio") or "").strip(),
            "problem_description": str(row.get("Descripcion_Problema") or "").strip(),
            "current_state": str(row.get("Estado_Actual") or "").strip(),
            "target_state": str(row.get("Estado_Objetivo") or "").strip(),
            "benefit": str(row.get("Beneficio") or "").strip(),
            "business_criticality_raw": str(row.get("Criticidad_Negocio") or "").strip(),
            "business_criticality": criticality_label,
            "business_criticality_rank": criticality_rank,
            "risk_category_raw": str(row.get("Categoria_Riesgo") or "").strip(),
            "priority_theme": theme,
            "priority_theme_rank": THEME_RANK[theme],
            "priority_theme_confidence": confidence,
            "needs_review": needs_review,
            "theme_scores": theme_scores,
            "dimension": dimension or "",
            "sprint_raw": str(row.get("Sprint") or "").strip(),
            "increment_raw": str(row.get("Incremento") or "").strip(),
            "period_raw": str(row.get("Periodo") or "").strip(),
            "sprint_hint": normalize_sprint_hint(row.get("Sprint"), row.get("Incremento"), row.get("Periodo")),
            "effort_level": str(row.get("Esfuerzo_Nivel") or "").strip(),
            "effort_hours_max": coerce_int(row.get("Esfuerzo_Horas_Max")),
            "responsable": str(row.get("Responsable") or "").strip(),
            "source_sheet": source_sheet,
            "source_type": "excel-backlog",
            "source_path": f"{BACKLOG_XLSX.name}::{source_sheet}",
            "vss_refs": pick_vss_refs(repo),
            "phase2_refs": pick_phase2_refs(repo),
        }
        records.append(raw_record)

    records, blocks = assign_group_sprints(records)

    total_records = len(records)
    critical_high_count = sum(1 for record in records if record["business_criticality_rank"] >= 3)
    unique_value_streams = sorted({record["value_stream_group"] for record in records})

    sprint_totals: Counter[str] = Counter(record["sprint_assigned"] for record in records)
    coverage: list[dict[str, object]] = []
    cumulative = 0
    for sprint in SPRINTS:
        count = sprint_totals[sprint.sprint_id]
        cumulative += count
        coverage.append(
            {
                "sprint_id": sprint.sprint_id,
                "label": sprint.label,
                "start": sprint.start.isoformat(),
                "end": sprint.end.isoformat(),
                "count": count,
                "cumulative_count": cumulative,
                "percent": round((count / total_records) * 100, 2) if total_records else 0,
                "cumulative_percent": round((cumulative / total_records) * 100, 2) if total_records else 0,
            }
        )

    stack_summaries: list[dict[str, object]] = []
    for stack in STACK_ORDER:
        stack_records = [record for record in records if record["stack"] == stack]
        if not stack_records:
            continue
        theme_counts = Counter(record["priority_theme"] for record in stack_records)
        repo_counts = Counter(record["repo"] for record in stack_records)
        value_stream_counts = Counter(record["value_stream_group"] for record in stack_records)
        stack_summaries.append(
            {
                "stack": stack,
                "repo_count": len({record["repo"] for record in stack_records}),
                "record_count": len(stack_records),
                "critical_high_count": sum(1 for record in stack_records if record["business_criticality_rank"] >= 3),
                "severity_score": sum(record["business_criticality_rank"] for record in stack_records),
                "top_themes": theme_counts.most_common(3),
                "top_repos": repo_counts.most_common(4),
                "top_value_streams": value_stream_counts.most_common(4),
                "sample_pain_points": [record["pain_point"] for record in sorted(stack_records, key=lambda item: (-item["business_criticality_rank"], item["priority_theme_rank"]))[:5]],
            }
        )

    theme_summaries: list[dict[str, object]] = []
    for theme in THEME_ORDER:
        theme_records = [record for record in records if record["priority_theme"] == theme]
        if not theme_records:
            continue
        stack_counts = Counter(record["stack"] for record in theme_records)
        value_stream_counts = Counter(record["value_stream_group"] for record in theme_records)
        theme_summaries.append(
            {
                "theme": theme,
                "rank": THEME_RANK[theme],
                "record_count": len(theme_records),
                "critical_high_count": sum(1 for record in theme_records if record["business_criticality_rank"] >= 3),
                "top_stacks": stack_counts.most_common(4),
                "top_value_streams": value_stream_counts.most_common(5),
                "sample_pain_points": [record["pain_point"] for record in sorted(theme_records, key=lambda item: (-item["business_criticality_rank"], item["stack"]))[:5]],
                "color": PRIORITY_COLORS[theme],
            }
        )

    value_stream_summaries = []
    value_stream_records: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        value_stream_records[record["value_stream_group"]].append(record)
    for stream_name, stream_records in sorted(
        value_stream_records.items(),
        key=lambda item: (-len(item[1]), item[0]),
    ):
        value_stream_summaries.append(
            {
                "name": stream_name,
                "record_count": len(stream_records),
                "stacks": sorted({record["stack"] for record in stream_records}),
                "themes": Counter(record["priority_theme"] for record in stream_records).most_common(3),
                "sample_pain_points": [record["pain_point"] for record in sorted(stream_records, key=lambda item: (-item["business_criticality_rank"], item["priority_theme_rank"]))[:3]],
            }
        )

    general_plan = build_general_plan()

    sources = {
        "backlog": str(BACKLOG_XLSX),
        "phase2_zip": str(PHASE2_ZIP),
        "vss_root": str(VSS_ROOT),
    }

    return {
        "metadata": {
            "title": "Roadmap Ejecutivo de Remediación por Stack y VSS",
            "generated_at": date.today().isoformat(),
            "repo_count": len({record["repo"] for record in records}),
            "stack_count": len({record["stack"] for record in records}),
            "kickoff_start": KICKOFF_DATE.isoformat(),
            "horizon_start": SPRINTS[0].start.isoformat(),
            "horizon_end": SPRINTS[-1].end.isoformat(),
            "sprint_length_days": 10,
            "cadence": "quincenal laboral ajustada por asuetos",
            "current_focus": "Arranque operativo en S31 desde 17-mar-2026 y redistribucion de trabajo estructural desde S34",
            "sources": sources,
        },
        "priorities": {
            "order": THEME_ORDER,
            "colors": PRIORITY_COLORS,
            "windows": THEME_WINDOWS,
        },
        "sprints": [
            {
                "id": sprint.sprint_id,
                "label": sprint.label,
                "start": sprint.start.isoformat(),
                "end": sprint.end.isoformat(),
            }
            for sprint in SPRINTS
        ],
        "stats": {
            "total_records": total_records,
            "critical_high_count": critical_high_count,
            "stack_count": len({record["stack"] for record in records}),
            "repo_count": len({record["repo"] for record in records}),
            "value_stream_count": len(unique_value_streams),
            "needs_review_count": sum(1 for record in records if record["needs_review"]),
            "tracking_action_count": len(tracking["actions"]),
            "tracking_repo_count": len({action["repo"] for action in tracking["actions"]}),
            "tracking_pending_count": sum(1 for action in tracking["actions"] if action["status"] == "PENDIENTE"),
            "coverage_final_percent": 100.0 if total_records else 0,
        },
        "filters": {
            "stacks": STACK_ORDER,
            "themes": THEME_ORDER,
            "repos": sorted({record["repo"] for record in records}),
            "sprints": [sprint.sprint_id for sprint in SPRINTS],
            "value_streams": unique_value_streams,
        },
        "dimension_catalog": dimension_catalog,
        "tracking": tracking,
        "coverage": coverage,
        "general_plan": general_plan,
        "stack_summaries": stack_summaries,
        "theme_summaries": theme_summaries,
        "value_stream_summaries": value_stream_summaries,
        "gantt_blocks": [
            {
                "stack": block["stack"],
                "theme": block["theme"],
                "theme_rank": THEME_RANK[block["theme"]],
                "value_stream_group": block["value_stream_group"],
                "record_count": block["record_count"],
                "critical_high_count": block["critical_high_count"],
                "severity_score": block["severity_score"],
                "repos": block["repos"],
                "value_streams": block["value_streams"],
                "sample_pain_points": block["sample_pain_points"],
                "sprint_id": block["sprint_id"],
                "sprint_label": block["sprint_label"],
                "sprint_start": block["sprint_start"],
                "sprint_end": block["sprint_end"],
            }
            for block in blocks
        ],
        "records": records,
    }


def write_outputs(dataset: dict[str, object]) -> None:
    OUTPUT_JSON.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.__ROADMAP_EXEC_DATA__ = " + json.dumps(dataset, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def main() -> None:
    dataset = build_dataset()
    write_outputs(dataset)
    print(f"Generated {OUTPUT_JSON.name} and {OUTPUT_JS.name}")
    print(
        f"Repos: {dataset['stats']['repo_count']} | "
        f"Pain points: {dataset['stats']['total_records']} | "
        f"Needs review: {dataset['stats']['needs_review_count']}"
    )


if __name__ == "__main__":
    main()
